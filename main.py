from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from datetime import datetime
import openpyxl

# Authenticate with Google Drive
gauth = GoogleAuth() 
# This creates a local webserver and automatically handles authentication.
gauth.LocalWebserverAuth() 
# Create a GoogleDrive instance
drive = GoogleDrive(gauth) 

# Define the ID of the root folder you want to start from
root_folder_id = "YOUR_ROOT_FOLDER_ID_HERE"

# Create a new Excel workbook and add a worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the column headers
ws.cell(row=1, column=1, value="Folder Name")
ws.cell(row=1, column=2, value="File Name")
ws.cell(row=1, column=3, value="Modified Date")

# Initialize the row counter
current_row = 2

# Function to recursively list all subfolders and files and add them to the Excel sheet
def list_folders_and_files(folder_id, parent_path=""):
    
    global current_row

    folder = drive.CreateFile({'id': folder_id})
    folder.FetchMetadata(fields="title")
    folder_name = folder["title"]

    # List all files and folders in the current folder
    file_list = drive.ListFile({'q': f"'{folder_id}' in parents and trashed=false"}).GetList()
    
    for file_or_folder in file_list:
        if file_or_folder["mimeType"] == "application/vnd.google-apps.folder":
            # Recursively list subfolders and files in this folder
            list_folders_and_files(file_or_folder["id"], parent_path + "/" + folder_name)
        else:
            # Get the file name and modified date
            file_name = file_or_folder["title"]
            modified_date = datetime.strptime(file_or_folder["modifiedDate"], '%Y-%m-%dT%H:%M:%S.%fZ')
            formatted_date = modified_date.strftime('%d-%m-%Y')
            
            # Add data to the Excel sheet
            ws.cell(row=current_row, column=1, value=folder_name)
            ws.cell(row=current_row, column=2, value=file_name)
            ws.cell(row=current_row, column=3, value=formatted_date)
            current_row += 1
            print(folder_name + ' / ' + file_name)
    

# Start listing from the root folder
list_folders_and_files(root_folder_id)

# Save the Excel file
wb.save("file_list.xlsx")

# Upload the file to your Google Drive
upload_file = drive.CreateFile()
upload_file.SetContentFile('file_list.xlsx')
upload_file.Upload()
print('title: %s, mimeType: %s' % (upload_file['title'], upload_file['mimeType']))
